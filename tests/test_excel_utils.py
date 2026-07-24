"""Tests for Excel list maintenance helpers."""

from __future__ import annotations

from pathlib import Path
import shutil
import unittest
from unittest.mock import Mock, patch

from openpyxl import Workbook

from picorgftp_sql import excel_utils


def _workspace_temp(name: str) -> Path:
    root = Path(__file__).resolve().parents[1] / "tmp_test" / name
    if root.exists():
        shutil.rmtree(root)
    root.mkdir(parents=True)
    return root


def _write_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in ["NAZWY", "TYPY", "MODELE", "KOLORY", "DODATKI"]:
        workbook.create_sheet(sheet_name)
    entries = workbook.create_sheet("ENTRIES")
    entries.append(excel_utils.ENTRY_HEADERS)
    entries.append(
        [
            "5901234567890",
            "MAGGIORE",
            "KOMODA",
            "MA03",
            "BIALY",
            "DAB",
            "",
            "LED-RGB",
            "PRD-1",
        ]
    )
    workbook.save(path)


class ExcelUtilsTests(unittest.TestCase):
    def test_find_list_value_usage_dispatches_to_active_sqlite_store(self) -> None:
        adapter = Mock()
        adapter.find_list_value_usage.return_value = [{"product_id": "PRD-1"}]

        with patch.object(excel_utils, "_active_sqlite_store", return_value=adapter):
            result = excel_utils.find_list_value_usage("NAZWY", "MAGGIORE")

        self.assertEqual(result, [{"product_id": "PRD-1"}])
        adapter.find_list_value_usage.assert_called_once_with(
            "NAZWY", "MAGGIORE", limit=100
        )

    def test_find_list_value_usage_reads_entries_from_workbook(self) -> None:
        temp_dir = _workspace_temp("excel_usage")
        try:
            workbook_path = temp_dir / "lists.xlsx"
            _write_workbook(workbook_path)

            with patch.object(excel_utils.settings, "LISTS_WORKBOOK_PATH", str(workbook_path)):
                usage = excel_utils.find_list_value_usage("KOLORY", "bialy")

            self.assertEqual(len(usage), 1)
            self.assertEqual(usage[0]["product_id"], "PRD-1")
            self.assertEqual(usage[0]["fields"], "KOLOR1")
            self.assertIn("MAGGIORE", usage[0]["label"])
        finally:
            shutil.rmtree(temp_dir)

    def test_find_list_value_usage_normalizes_extra_underscores(self) -> None:
        temp_dir = _workspace_temp("excel_usage_extra")
        try:
            workbook_path = temp_dir / "lists.xlsx"
            _write_workbook(workbook_path)

            with patch.object(excel_utils.settings, "LISTS_WORKBOOK_PATH", str(workbook_path)):
                usage = excel_utils.find_list_value_usage("DODATKI", "led_rgb")

            self.assertEqual(len(usage), 1)
            self.assertEqual(usage[0]["fields"], "DODATKI")
        finally:
            shutil.rmtree(temp_dir)

    def test_find_list_value_usage_normalizes_polish_l_stroke_in_legacy_workbook(
        self,
    ) -> None:
        temp_dir = _workspace_temp("excel_usage_polish_l_stroke")
        try:
            workbook_path = temp_dir / "lists.xlsx"
            _write_workbook(workbook_path)
            workbook = excel_utils.load_workbook(workbook_path)
            workbook["ENTRIES"].cell(row=2, column=3).value = "STÓŁ"
            workbook.save(workbook_path)
            workbook.close()

            with patch.object(excel_utils.settings, "LISTS_WORKBOOK_PATH", str(workbook_path)):
                usage = excel_utils.find_list_value_usage("TYPY", "stol")

            self.assertEqual(len(usage), 1)
            self.assertEqual(usage[0]["fields"], "TYP")
        finally:
            shutil.rmtree(temp_dir)


if __name__ == "__main__":
    unittest.main()
