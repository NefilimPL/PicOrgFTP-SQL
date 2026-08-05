"""Tests for importing legacy files into one SQLite database."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from picorgftp_sql import config, sqlite_store
from picorgftp_sql.legacy_import import import_legacy_to_sqlite
from picorgftp_sql.sqlite_store import (
    ProductEanConflictError,
    SqliteStore,
)

ProductIdConflictError = getattr(sqlite_store, "ProductIdConflictError", ValueError)


def _write_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, value in {
        "NAZWY": "MAGGIORE",
        "TYPY": "KOMODA",
        "MODELE": "MA03",
        "KOLORY": "BIALY",
        "DODATKI": "NO-LED",
    }.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append([value])
    entries = workbook.create_sheet("ENTRIES")
    entries.append(
        [
            "EAN",
            "NAZWA",
            "TYP",
            "MODEL",
            "KOLOR1",
            "KOLOR2",
            "KOLOR3",
            "DODATKI",
            "PRODUCT_ID",
        ]
    )
    entries.append(
        [
            "5901234567890",
            "MAGGIORE",
            "KOMODA",
            "MA03",
            "BIALY",
            "",
            "",
            "NO-LED",
            "PRD-1",
        ]
    )
    workbook.save(path)


def test_import_legacy_files_to_sqlite(tmp_path: Path) -> None:
    legacy_dir = tmp_path / "legacy"
    legacy_dir.mkdir()
    raw_config = json.loads(json.dumps(config.DEFAULT_CONFIG))
    raw_config[config.p] = "mysql"
    raw_config[config.SQL_AVAILABLE_COLUMNS_KEY] = ["img_01"]
    (legacy_dir / "config.json").write_text(
        json.dumps(raw_config), encoding="utf-8"
    )
    _write_workbook(legacy_dir / "lists.xlsx")
    (legacy_dir / "web_users.json").write_text(
        json.dumps(
            [
                {
                    "username": "operator",
                    "role": "user",
                    "enabled": True,
                    "password_hash": "hash",
                }
            ]
        ),
        encoding="utf-8",
    )
    (legacy_dir / "web_history.json").write_text(
        json.dumps(
            [
                {
                    "id": "hist-1",
                    "ts": 1.0,
                    "time": "2026-06-25 12:00:00",
                    "user": "operator",
                    "action": "save",
                    "ean": "5901234567890",
                    "details": {},
                }
            ]
        ),
        encoding="utf-8",
    )
    (legacy_dir / "file_index.json").write_text(
        json.dumps({"version": 1, "names": ["MAGGIORE"]}),
        encoding="utf-8",
    )

    result = import_legacy_to_sqlite(
        legacy_dir=str(legacy_dir),
        database_path=str(tmp_path / "data.sqlite"),
    )
    store = SqliteStore(str(tmp_path / "data.sqlite"))

    assert result["ok"] is True
    assert result["config"] is True
    assert result["entries"] == 1
    assert result["users"] == 1
    assert result["history"] == 1
    assert result["file_index"] is True
    assert store.load_config()[config.p] == "mysql"
    assert store.load_sql_columns() == ["img_01"]
    assert store.load_lists()["__ENTRY_RECORDS__"][0]["PRODUCT_ID"] == "PRD-1"
    assert store.load_users()[0]["username"] == "operator"
    assert store.load_history()[0]["id"] == "hist-1"
    imported_index = store.load_file_index_cache()
    assert imported_index["version"] == 1
    assert imported_index["names"] == ["MAGGIORE"]
    assert isinstance(imported_index["generated_at"], str)
    assert imported_index["generated_at"].endswith("Z")
    assert "T" in imported_index["generated_at"]


@pytest.mark.parametrize(
    ("duplicate_row", "conflict_error"),
    (
        (
            [" 5901234567890 ", "DUPLICATE", "KOMODA", "MA03", "BIALY", "", "", "NO-LED", ""],
            ProductEanConflictError,
        ),
        (
            ["5901234567891", "DUPLICATE", "KOMODA", "MA03", "BIALY", "", "", "NO-LED", " prd-1 "],
            ProductIdConflictError,
        ),
    ),
)
def test_conflicting_legacy_identity_is_rejected_before_target_mutation(
    tmp_path: Path,
    duplicate_row,
    conflict_error,
) -> None:
    """Invalid workbook identity data must not leave a partially imported target."""

    legacy_dir = tmp_path / "legacy"
    legacy_dir.mkdir()
    source_config = json.loads(json.dumps(config.DEFAULT_CONFIG))
    source_config[config.p] = "mysql"
    (legacy_dir / "config.json").write_text(
        json.dumps(source_config), encoding="utf-8"
    )
    workbook_path = legacy_dir / "lists.xlsx"
    _write_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    workbook["ENTRIES"].append(duplicate_row)
    workbook.save(workbook_path)
    workbook.close()

    db_path = tmp_path / "data.sqlite"
    target = SqliteStore(str(db_path))
    target.save_config({config.p: "before-import"})
    target.save_product_entry(
        {"PRODUCT_ID": "P-OLD", "EAN": "OLD-EAN", "NAZWA": "EXISTING"}
    )

    with pytest.raises(conflict_error):
        import_legacy_to_sqlite(str(legacy_dir), str(db_path))

    assert target.load_config()[config.p] == "before-import"
    assert target.get_product_by_id("P-OLD")["NAZWA"] == "EXISTING"
