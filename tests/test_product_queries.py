"""Tests for the common product query contract."""

import json
from pathlib import Path
import statistics
import time
from unittest.mock import Mock

from openpyxl import Workbook
import pytest

from picorgftp_sql.product_queries import (
    ProductSearchCriteria,
    filter_product_records,
)
from picorgftp_sql import data_store, excel_utils, web_data
from picorgftp_sql.sqlite_store import SqliteStore


RECORDS = [
    {"PRODUCT_ID": "P-1", "EAN": "5901", "NAZWA": "ALFA", "TYP": "STÓŁ", "MODEL": "A1"},
    {"PRODUCT_ID": "P-2", "EAN": "5902", "NAZWA": "BETA", "TYP": "SZAFA", "MODEL": "B1"},
]

_BENCHMARK_PRODUCT_COUNT = 100_000
_BENCHMARK_INSERT_SQL = """
    INSERT INTO product_entries (
        product_id, ean, name, type_name, model,
        product_id_key, ean_key, name_key, type_name_key, model_key,
        color1, color2, color3, extra,
        color1_key, color2_key, color3_key, extra_key,
        search_text_key, updated_at
    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
"""


def benchmark_product_rows(count: int = _BENCHMARK_PRODUCT_COUNT):
    """Yield repeatable, workbook-free product rows for SQLite benchmarks."""

    for index in range(count):
        product_id = f"PRD-{index:06d}"
        ean = f"5901{index:09d}"
        name = f"PRODUKT-{index % 250:03d}"
        type_name = f"TYP-{index % 25:02d}"
        model = f"MODEL-{index % 100:03d}"
        color1 = ("BIALY", "CZARNY", "DAB")[index % 3]
        color2 = ""
        color3 = ""
        extra = ("NO-LED", "LED")[index % 2]
        keys = tuple(
            value.casefold()
            for value in (
                product_id,
                ean,
                name,
                type_name,
                model,
                color1,
                color2,
                color3,
                extra,
            )
        )
        yield (
            product_id,
            ean,
            name,
            type_name,
            model,
            *keys[:5],
            color1,
            color2,
            color3,
            extra,
            *keys[5:],
            " ".join(keys),
            "2026-07-27T00:00:00.000Z",
        )


def write_product_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in ("NAZWY", "TYPY", "MODELE", "KOLORY", "DODATKI"):
        workbook.create_sheet(sheet_name)
    entries = workbook.create_sheet("ENTRIES")
    entries.append(excel_utils.ENTRY_HEADERS)
    entries.append(
        [
            "5901234567890",
            "ALFA",
            "STÓŁ",
            "A1",
            "BIAŁY",
            "",
            "",
            "NO-LED",
            "P-1",
        ]
    )
    workbook.save(path)


def test_prepare_excel_lists_reuses_snapshot_until_mtime_changes(monkeypatch, tmp_path):
    """Catch a regression that reloads an unchanged legacy workbook."""

    workbook_path = tmp_path / "data.xlsx"
    write_product_workbook(workbook_path)
    loads = 0
    original_load_workbook = excel_utils.load_workbook

    def fake_load_workbook(*args, **kwargs):
        nonlocal loads
        loads += 1
        return original_load_workbook(*args, **kwargs)

    monkeypatch.setattr(excel_utils, "load_workbook", fake_load_workbook)
    monkeypatch.setattr(excel_utils.settings, "LISTS_WORKBOOK_PATH", str(workbook_path))
    excel_utils.clear_excel_snapshot_cache()

    excel_utils.prepare_excel_lists()
    excel_utils.prepare_excel_lists()

    assert loads == 1


def test_prepare_excel_lists_returns_independent_snapshot_copies(monkeypatch, tmp_path):
    """Catch cached list or entry data leaking mutations between callers."""

    workbook_path = tmp_path / "data.xlsx"
    write_product_workbook(workbook_path)
    monkeypatch.setattr(excel_utils.settings, "LISTS_WORKBOOK_PATH", str(workbook_path))
    excel_utils.clear_excel_snapshot_cache()

    first = excel_utils.prepare_excel_lists()
    first["NAZWY"].append("MUTATED")
    first["ENTRIES"]["5901234567890"]["NAZWA"] = "MUTATED"
    first[excel_utils.ENTRY_RECORDS_KEY][0]["NAZWA"] = "MUTATED"

    second = excel_utils.prepare_excel_lists()

    assert second["NAZWY"] == []
    assert second["ENTRIES"]["5901234567890"]["NAZWA"] == "ALFA"
    assert second[excel_utils.ENTRY_RECORDS_KEY][0]["NAZWA"] == "ALFA"


def test_failed_workbook_save_keeps_existing_excel_snapshot(monkeypatch, tmp_path):
    """Catch cache eviction when an application workbook save fails."""

    workbook_path = tmp_path / "data.xlsx"
    write_product_workbook(workbook_path)
    loads = 0
    original_load_workbook = excel_utils.load_workbook

    def fake_load_workbook(*args, **kwargs):
        nonlocal loads
        loads += 1
        return original_load_workbook(*args, **kwargs)

    class FailingWorkbook:
        def save(self, _path):
            raise OSError("write failed")

    monkeypatch.setattr(excel_utils, "load_workbook", fake_load_workbook)
    monkeypatch.setattr(excel_utils.settings, "LISTS_WORKBOOK_PATH", str(workbook_path))
    monkeypatch.setattr(excel_utils.messagebox, "showerror", lambda *_args: None)
    excel_utils.clear_excel_snapshot_cache()

    before_save = excel_utils.prepare_excel_lists()

    assert not excel_utils._save_workbook(FailingWorkbook(), "test_save_failed")
    assert excel_utils.prepare_excel_lists() == before_save
    assert loads == 1


def test_successful_workbook_save_clears_existing_excel_snapshot(monkeypatch, tmp_path):
    """Catch a successful application save that leaves stale data cached."""

    workbook_path = tmp_path / "data.xlsx"
    write_product_workbook(workbook_path)
    loads = 0
    original_load_workbook = excel_utils.load_workbook

    def fake_load_workbook(*args, **kwargs):
        nonlocal loads
        loads += 1
        return original_load_workbook(*args, **kwargs)

    class SuccessfulWorkbook:
        def save(self, _path):
            return None

    monkeypatch.setattr(excel_utils, "load_workbook", fake_load_workbook)
    monkeypatch.setattr(excel_utils.settings, "LISTS_WORKBOOK_PATH", str(workbook_path))
    excel_utils.clear_excel_snapshot_cache()

    excel_utils.prepare_excel_lists()

    assert excel_utils._save_workbook(SuccessfulWorkbook(), "test_save_succeeded")
    excel_utils.prepare_excel_lists()

    assert loads == 2


def test_filter_product_records_prefers_exact_identity_and_limits():
    criteria = ProductSearchCriteria(product_id="p-1")

    assert filter_product_records(RECORDS, criteria, limit=1) == [RECORDS[0]]


def test_legacy_store_uses_prepared_records_for_exact_identity_lookups(monkeypatch):
    monkeypatch.setattr(
        excel_utils,
        "prepare_excel_lists",
        lambda: {excel_utils.ENTRY_RECORDS_KEY: RECORDS},
    )
    store = data_store.LegacyDataStore()

    assert store.get_product_by_id(" p-1 ") == RECORDS[0]
    assert store.get_product_by_ean("5902") == RECORDS[1]


def test_legacy_store_suggests_values_from_prepared_records(monkeypatch):
    monkeypatch.setattr(
        excel_utils,
        "prepare_excel_lists",
        lambda: {excel_utils.ENTRY_RECORDS_KEY: RECORDS},
    )

    assert data_store.LegacyDataStore().suggest_product_field(
        "model", "a", {"name": "ALFA"}, limit=1
    ) == ["A1"]


def test_sqlite_product_queries_do_not_load_all_lists(tmp_path, monkeypatch):
    """Catch a regression to the full SQLite product-table materialization path."""

    adapter = data_store.SqliteDataStoreAdapter(str(tmp_path / "products.sqlite"))
    adapter.save_product_entry(
        {
            "PRODUCT_ID": "P-1",
            "EAN": "5901234567890",
            "NAZWA": "ALFA",
            "TYP": "STÓŁ",
            "MODEL": "A1",
        }
    )
    monkeypatch.setattr(
        adapter.store,
        "load_lists",
        lambda: (_ for _ in ()).throw(AssertionError("full load")),
    )

    expected = {
        "PRODUCT_ID": "P-1",
        "EAN": "5901234567890",
        "NAZWA": "ALFA",
        "TYP": "STÓŁ",
        "MODEL": "A1",
        "KOLOR1": "",
        "KOLOR2": "",
        "KOLOR3": "",
        "DODATKI": "NO-LED",
    }
    legacy = data_store.LegacyDataStore()
    monkeypatch.setattr(legacy, "_product_records", lambda: [expected])
    assert adapter.get_product_by_ean("5901234567890") == legacy.get_product_by_ean(
        "5901234567890"
    )
    assert adapter.get_product_by_id("p-1") == legacy.get_product_by_id("p-1")
    criteria = ProductSearchCriteria(name="alfa")
    assert adapter.search_product_entries(criteria, limit=1) == legacy.search_product_entries(
        criteria, limit=1
    )
    assert adapter.suggest_product_field(
        "model", "a", {"name": "ALFA"}
    ) == legacy.suggest_product_field("model", "a", {"name": "ALFA"})


def test_search_entries_delegates_to_active_store(monkeypatch):
    """Catch a regression to web-side product-list materialization."""

    store = Mock()
    store.search_product_entries.return_value = [
        {"PRODUCT_ID": "P-1", "EAN": "5901", "NAZWA": "ALFA"}
    ]
    monkeypatch.setattr(web_data, "get_active_store", lambda: store)
    monkeypatch.setattr(
        web_data,
        "prepare_excel_lists",
        lambda: (_ for _ in ()).throw(AssertionError("full load")),
    )

    result = web_data.search_entries(ean="5901", limit=10)

    assert result[0]["product_id"] == "P-1"
    store.search_product_entries.assert_called_once_with(
        ProductSearchCriteria(ean="5901"), limit=10
    )


def test_web_search_query_is_filtered_by_sqlite_store_before_limit(tmp_path, monkeypatch):
    """A matching row after an earlier criterion match must remain reachable."""

    store = data_store.SqliteDataStoreAdapter(str(tmp_path / "products.sqlite"))
    store.save_product_entry(
        {
            "PRODUCT_ID": "P-1",
            "EAN": "5901",
            "NAZWA": "ALFA",
            "TYP": "STOL",
            "MODEL": "A1",
            "DODATKI": "MISS",
        }
    )
    store.save_product_entry(
        {
            "PRODUCT_ID": "P-2",
            "EAN": "5902",
            "NAZWA": "ALFA",
            "TYP": "STOL",
            "MODEL": "A2",
            "DODATKI": "TARGET",
        }
    )
    monkeypatch.setattr(web_data, "get_active_store", lambda: store)
    monkeypatch.setattr(
        web_data,
        "prepare_excel_lists",
        lambda: (_ for _ in ()).throw(AssertionError("full load")),
    )

    result = web_data.search_entries(name="ALFA", query="target", limit=1)

    assert [entry["product_id"] for entry in result] == ["P-2"]


def test_search_entries_clamps_zero_and_negative_limits(monkeypatch):
    """Store delegation must keep non-positive product limits bounded."""

    store = Mock()
    store.search_product_entries.return_value = []
    monkeypatch.setattr(web_data, "get_active_store", lambda: store)

    web_data.search_entries(limit=0)
    web_data.search_entries(limit=-1)

    assert [call.kwargs["limit"] for call in store.search_product_entries.call_args_list] == [
        1,
        1,
    ]


def test_find_entry_by_ean_uses_store_and_returns_browser_payload(monkeypatch):
    """Identity lookup must retain the lowercase browser response contract."""

    store = Mock()
    store.get_product_by_ean.return_value = {
        "PRODUCT_ID": "P-1",
        "EAN": "5901",
        "NAZWA": "ALFA",
        "TYP": "STOL",
        "MODEL": "A1",
        "KOLOR1": "BIALY",
        "KOLOR2": "",
        "KOLOR3": "",
        "DODATKI": "NO-LED",
    }
    monkeypatch.setattr(web_data, "get_active_store", lambda: store)

    result = web_data.find_entry_by_identity(ean="5901")

    assert result is not None
    assert result["product_id"] == "P-1"
    assert result["ean"] == "5901"
    assert result["extra"] == "NO-LED"
    store.get_product_by_ean.assert_called_once_with("5901")


def test_field_suggestions_delegates_to_active_store(monkeypatch):
    """Catch a regression to loading all product rows for suggestions."""

    store = Mock()
    store.mode = "sqlite"
    store.suggest_product_field.return_value = ["ALFA"]
    monkeypatch.setattr(web_data, "get_active_store", lambda: store)
    monkeypatch.setattr(web_data, "_get_file_index", lambda **_kwargs: None)
    monkeypatch.setattr(
        web_data,
        "prepare_excel_lists",
        lambda: (_ for _ in ()).throw(AssertionError("full load")),
    )

    result = web_data.field_suggestions("name", {"name": "al"}, limit=10)

    assert result == ["ALFA"]
    store.suggest_product_field.assert_called_once_with(
        "name",
        "al",
        {
            "product_id": "",
            "ean": "",
            "name": "",
            "type_name": "",
            "model": "",
        },
        limit=10,
    )


def test_save_web_entry_uses_one_active_store_identity_lookup(monkeypatch):
    """Catch duplicate product-ID and EAN pre-lookups before a save."""

    store = Mock()
    store.get_product_by_id.return_value = {
        "PRODUCT_ID": "P-1",
        "EAN": "5901",
    }
    monkeypatch.setattr(web_data, "get_active_store", lambda: store)
    save_entry = Mock(return_value={"updated": True, "product_id": "P-1", "entry": {}})
    monkeypatch.setattr(web_data, "save_ean_entry", save_entry)

    web_data.save_web_entry(
        {
            "product_id": "P-1",
            "ean": "5901",
            "name": "ALFA",
            "type_name": "STOL",
            "model": "A1",
            "color1": "BIALY",
        }
    )

    store.get_product_by_id.assert_called_once_with("P-1")
    store.get_product_by_ean.assert_not_called()


@pytest.mark.performance
def test_100000_product_selective_query_benchmark(tmp_path):
    """Keep indexed identity lookups and bounded suggestions within budgets."""

    store = SqliteStore(str(tmp_path / "products-100000.sqlite"))
    store.initialize()
    with store.connection() as conn:
        conn.executemany(
            _BENCHMARK_INSERT_SQL,
            benchmark_product_rows(),
        )

    target_index = _BENCHMARK_PRODUCT_COUNT - 1
    target_id = f"PRD-{target_index:06d}"
    target_ean = f"5901{target_index:09d}"

    # Warm the database page cache and each production query path before sampling.
    assert store.get_product_by_id(target_id)["EAN"] == target_ean
    assert store.get_product_by_ean(target_ean)["PRODUCT_ID"] == target_id
    assert store.suggest_product_field("name", "produkt-", {}, limit=50)

    lookup_samples = []
    for index in range(200):
        started = time.perf_counter()
        if index % 2:
            result = store.get_product_by_id(target_id)
        else:
            result = store.get_product_by_ean(target_ean)
        lookup_samples.append(time.perf_counter() - started)
        assert result is not None
        assert result["PRODUCT_ID"] == target_id

    suggestion_samples = []
    suggestions = []
    for _ in range(100):
        started = time.perf_counter()
        suggestions = store.suggest_product_field(
            "name", "produkt-", {}, limit=50
        )
        suggestion_samples.append(time.perf_counter() - started)

    lookup_p50 = statistics.median(lookup_samples)
    lookup_p95 = statistics.quantiles(lookup_samples, n=100)[94]
    suggestion_p50 = statistics.median(suggestion_samples)
    suggestion_p95 = statistics.quantiles(suggestion_samples, n=100)[94]

    with store.connection() as conn:
        ean_plan = [
            row["detail"]
            for row in conn.execute(
                "EXPLAIN QUERY PLAN SELECT product_id FROM product_entries "
                "WHERE ean_key = ? ORDER BY rowid LIMIT 1",
                (target_ean.casefold(),),
            )
        ]
        product_id_plan = [
            row["detail"]
            for row in conn.execute(
                "EXPLAIN QUERY PLAN SELECT ean FROM product_entries "
                "WHERE product_id_key = ? ORDER BY rowid LIMIT 1",
                (target_id.casefold(),),
            )
        ]
        suggestion_plan = [
            row["detail"]
            for row in conn.execute(
                "EXPLAIN QUERY PLAN SELECT DISTINCT name FROM product_entries "
                "WHERE name <> '' AND name_key >= ? AND name_key < ? "
                "ORDER BY name_key, name LIMIT ?",
                ("produkt-", "produkt-\U0010ffff", 50),
            )
        ]

    assert any("idx_product_entries_ean_key" in detail for detail in ean_plan)
    assert any(
        "idx_product_entries_product_id_key" in detail
        for detail in product_id_plan
    )
    assert any(
        "idx_product_entries_name_key" in detail for detail in suggestion_plan
    )
    assert not any(
        "SCAN product_entries" in detail
        for detail in (*ean_plan, *product_id_plan, *suggestion_plan)
    )

    report = {
        "lookup_p50_seconds": lookup_p50,
        "lookup_p95_seconds": lookup_p95,
        "lookup_samples": len(lookup_samples),
        "plans": {
            "ean": ean_plan,
            "product_id": product_id_plan,
            "suggestion": suggestion_plan,
        },
        "product_count": _BENCHMARK_PRODUCT_COUNT,
        "suggestion_count": len(suggestions),
        "suggestion_p50_seconds": suggestion_p50,
        "suggestion_p95_seconds": suggestion_p95,
        "suggestion_samples": len(suggestion_samples),
    }
    print(json.dumps(report, ensure_ascii=False, sort_keys=True))

    assert lookup_p95 < 0.050
    assert suggestion_p95 < 0.200
    assert len(suggestions) <= 50
