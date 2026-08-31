"""Import legacy JSON/Excel files into a SQLite data store."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .common import SQL_AVAILABLE_COLUMNS_KEY, SQL_COLUMN_MAP_KEY, SLOT_DEFS_KEY
from .excel_utils import (
    COLOR1_HEADER,
    COLOR2_HEADER,
    COLOR3_HEADER,
    EAN_HEADER,
    ENTRY_RECORDS_KEY,
    EXTRA_HEADER,
    MODEL_HEADER,
    NAME_HEADER,
    PRODUCT_ID_HEADER,
    TYPE_HEADER,
)
from .sqlite_store import LIST_SHEETS, SqliteStore

ENTRY_SHEET = "ENTRIES"
ENTRY_HEADERS = [
    EAN_HEADER,
    NAME_HEADER,
    TYPE_HEADER,
    MODEL_HEADER,
    COLOR1_HEADER,
    COLOR2_HEADER,
    COLOR3_HEADER,
    EXTRA_HEADER,
    PRODUCT_ID_HEADER,
]


def _read_json(path: Path, fallback: Any) -> Any:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError, TypeError):
        return fallback
    return payload


def _cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _header_map(row) -> dict[str, int]:
    mapping = {}
    for index, cell in enumerate(row or [], start=1):
        text = _cell(getattr(cell, "value", cell)).upper()
        if text:
            mapping[text] = index
    return mapping


def _row_value(row, mapping: dict[str, int], header: str) -> str:
    index = mapping.get(header)
    if not index or index > len(row):
        return ""
    return _cell(row[index - 1].value)


def _read_workbook_payload(path: Path) -> dict[str, Any]:
    payload: dict[str, Any] = {sheet: [] for sheet in LIST_SHEETS}
    payload[ENTRY_RECORDS_KEY] = []
    if not path.exists():
        return payload
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in LIST_SHEETS:
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            values = []
            for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
                value = _cell(row[0]).upper()
                if value and value not in values:
                    values.append(value)
            payload[sheet_name] = values
        if ENTRY_SHEET in workbook.sheetnames:
            sheet = workbook[ENTRY_SHEET]
            first_row = next(sheet.iter_rows(min_row=1, max_row=1), ())
            mapping = _header_map(first_row)
            records = []
            for row in sheet.iter_rows(min_row=2):
                ean = _row_value(row, mapping, EAN_HEADER)
                if not ean:
                    continue
                records.append(
                    {
                        EAN_HEADER: ean,
                        NAME_HEADER: _row_value(row, mapping, NAME_HEADER),
                        TYPE_HEADER: _row_value(row, mapping, TYPE_HEADER),
                        MODEL_HEADER: _row_value(row, mapping, MODEL_HEADER),
                        COLOR1_HEADER: _row_value(row, mapping, COLOR1_HEADER),
                        COLOR2_HEADER: _row_value(row, mapping, COLOR2_HEADER),
                        COLOR3_HEADER: _row_value(row, mapping, COLOR3_HEADER),
                        EXTRA_HEADER: _row_value(row, mapping, EXTRA_HEADER),
                        PRODUCT_ID_HEADER: _row_value(row, mapping, PRODUCT_ID_HEADER),
                    }
                )
            payload[ENTRY_RECORDS_KEY] = records
    finally:
        workbook.close()
    return payload


def _merge_mapping(existing: dict[str, Any], incoming: dict[str, Any]) -> dict[str, Any]:
    """Overlay legacy values while retaining keys held only by the SQLite source."""

    merged = dict(existing)
    for key, value in incoming.items():
        previous = merged.get(key)
        if isinstance(previous, dict) and isinstance(value, dict):
            merged[key] = _merge_mapping(previous, value)
        else:
            merged[key] = value
    return merged


def _merge_users(
    existing: list[dict[str, Any]], incoming: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    """Keep accounts present only in one old store; JSON wins on the same login."""

    users: list[dict[str, Any]] = []
    index_by_username: dict[str, int] = {}
    for item in [*existing, *incoming]:
        if not isinstance(item, dict):
            continue
        username = str(item.get("username") or "").strip()
        if not username:
            continue
        key = username.casefold()
        if key in index_by_username:
            users[index_by_username[key]] = dict(item)
        else:
            index_by_username[key] = len(users)
            users.append(dict(item))
    return users


def _entry_key(record: dict[str, Any]) -> tuple[str, str]:
    product_id = _cell(record.get(PRODUCT_ID_HEADER)).casefold()
    if product_id:
        return ("product_id", product_id)
    ean = _cell(record.get(EAN_HEADER)).casefold()
    return ("ean", ean) if ean else ("record", repr(sorted(record.items())))


def _merge_lists_payload(existing: dict[str, Any], incoming: dict[str, Any]) -> dict[str, Any]:
    """Combine old workbook values with entries already held in legacy SQLite."""

    merged: dict[str, Any] = {}
    for sheet in LIST_SHEETS:
        values: list[Any] = []
        seen: set[str] = set()
        for value in [*(existing.get(sheet, []) or []), *(incoming.get(sheet, []) or [])]:
            key = _cell(value).casefold()
            if key and key not in seen:
                seen.add(key)
                values.append(value)
        merged[sheet] = values

    records: list[dict[str, Any]] = []
    index_by_key: dict[tuple[str, str], int] = {}
    for item in [
        *(existing.get(ENTRY_RECORDS_KEY, []) or []),
        *(incoming.get(ENTRY_RECORDS_KEY, []) or []),
    ]:
        if not isinstance(item, dict):
            continue
        key = _entry_key(item)
        if key in index_by_key:
            records[index_by_key[key]] = dict(item)
        else:
            index_by_key[key] = len(records)
            records.append(dict(item))
    merged[ENTRY_RECORDS_KEY] = records
    return merged


def import_legacy_to_sqlite(
    legacy_dir: str,
    database_path: str,
    *,
    merge_existing: bool = False,
) -> dict[str, Any]:
    """Import supported legacy files from ``legacy_dir`` into ``database_path``."""

    source = Path(legacy_dir)
    config_path = source / "config.json"
    lists_path = source / "lists.xlsx"
    users_path = source / "web_users.json"
    history_path = source / "web_history.json"
    file_index_path = source / "file_index.json"
    raw_config = _read_json(config_path, {})
    config_imported = isinstance(raw_config, dict) and bool(raw_config)
    lists_imported = lists_path.is_file()
    lists_payload = _read_workbook_payload(lists_path) if lists_imported else {}
    users = _read_json(users_path, [])
    users_imported = users_path.is_file() and isinstance(users, list)
    if not users_imported:
        users = []
    history = _read_json(history_path, [])
    history_imported = history_path.is_file() and isinstance(history, list)
    if not history_imported:
        history = []
    file_index = _read_json(file_index_path, {})
    file_index_imported = isinstance(file_index, dict) and bool(file_index)

    store = SqliteStore(database_path)
    if lists_imported:
        store.validate_lists_payload(lists_payload)
    store.initialize()

    if config_imported:
        config_payload = (
            _merge_mapping(store.load_config(), raw_config)
            if merge_existing
            else raw_config
        )
        store.save_config(config_payload)
        columns = raw_config.get(SQL_AVAILABLE_COLUMNS_KEY, [])
        if isinstance(columns, list):
            store.save_sql_columns(columns)
        slot_defs = raw_config.get(SLOT_DEFS_KEY, [])
        sql_map = raw_config.get(SQL_COLUMN_MAP_KEY, {})
        if isinstance(slot_defs, list) and isinstance(sql_map, dict):
            store.save_slots(slot_defs, sql_map)

    if lists_imported:
        list_payload = (
            _merge_lists_payload(store.load_lists(), lists_payload)
            if merge_existing
            else lists_payload
        )
        store.save_lists(list_payload)

    if users_imported:
        incoming_users = [item for item in users if isinstance(item, dict)]
        store.save_users(
            _merge_users(store.load_users(), incoming_users)
            if merge_existing
            else incoming_users
        )

    if history_imported:
        incoming_history = [item for item in history if isinstance(item, dict)]
        store.save_history(
            [*store.load_history(), *incoming_history]
            if merge_existing
            else incoming_history
        )

    if file_index_imported:
        index_payload = (
            _merge_mapping(store.load_file_index_cache(), file_index)
            if merge_existing
            else file_index
        )
        store.save_file_index_cache(index_payload)

    records = lists_payload.get(ENTRY_RECORDS_KEY, [])
    return {
        "ok": True,
        "config": config_imported,
        "lists": sum(
            len(lists_payload.get(sheet, []))
            for sheet in LIST_SHEETS
            if isinstance(lists_payload.get(sheet, []), list)
        ),
        "entries": len(records) if isinstance(records, list) else 0,
        "users": len(users) if users_imported else 0,
        "history": len(history) if history_imported else 0,
        "file_index": file_index_imported,
    }
