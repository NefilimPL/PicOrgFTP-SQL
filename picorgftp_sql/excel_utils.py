"""Utilities for maintaining the Excel lists used by the application."""

from __future__ import annotations

import os
import uuid
from typing import Dict

from openpyxl import Workbook, load_workbook
from tkinter import messagebox

from .common import DEFAULT_SLOT_DEFS, ELEMENT_PIC, NON_PIC, OPEN_FURNITURE
from .logging_utils import log_error_loc, log_info_loc
from .system_utils import get_file_lock_user
from . import localization, settings

ENTRY_SHEET = "ENTRIES"
ENTRY_RECORDS_KEY = "__ENTRY_RECORDS__"
EAN_HEADER = "EAN"
NAME_HEADER = "NAZWA"
TYPE_HEADER = "TYP"
MODEL_HEADER = "MODEL"
COLOR1_HEADER = "KOLOR1"
COLOR2_HEADER = "KOLOR2"
COLOR3_HEADER = "KOLOR3"
EXTRA_HEADER = "DODATKI"
PRODUCT_ID_HEADER = "PRODUCT_ID"
EXTRAS_SHEET = "DODATKI"
ENTRY_HEADERS = [
    EAN_HEADER,
    NAME_HEADER,
    TYPE_HEADER,
    MODEL_HEADER,
    COLOR1_HEADER,
    COLOR2_HEADER,
    COLOR3_HEADER,
    EXTRA_HEADER,
    PRODUCT_ID_HEADER,
]

NO_EAN_PLACEHOLDER = "BRAK-EAN"
NO_LED_VALUE = "NO-LED"
EMPTY = ""
UNDERSCORE = "_"
HYPHEN = "-"
LOCKED_TITLE = localization.EXCEL_LOCKED_TITLE
LOCKED_REASON_OTHER_PROCESS = localization.EXCEL_LOCK_OTHER_PROCESS
LOCKED_BY_USER_TEMPLATE = localization.EXCEL_LOCKED_BY_USER
ERROR_TITLE = localization.AK
WRITE_ERROR_TITLE = localization.Ac

# Order used by the GUI when building slot labels.
SLOT_LABELS = [(slot["prefix"], slot["label"]) for slot in DEFAULT_SLOT_DEFS]


def _workbook_path() -> str:
    return settings.LISTS_WORKBOOK_PATH


def _excel_sheets() -> dict[str, str]:
    return settings.EXCEL_SHEETS


def label_category(label: str) -> str:
    """Return a human-readable category for a slot label."""

    base = label.rstrip("0123456789")
    if base.startswith("LED_"):
        base = base[4:]
    lowered = base.lower()
    if "assembly_instruction" in lowered:
        return "ASSEMBLY"
    if "technical_drawing" in lowered:
        return "TECHNICAL"
    if "mood_pic" in lowered:
        return "MOOD"
    if "wb_pic" in lowered:
        return "WB"
    if "detail_pic" in lowered:
        return "DETAIL"
    if ELEMENT_PIC in lowered:
        return "ELEMENT"
    if OPEN_FURNITURE in lowered:
        return "OPEN-FURNITURE"
    if "no_ean" in lowered:
        return "NO-EAN"
    if NON_PIC in lowered:
        return "NON-PIC"
    return base.replace(UNDERSCORE, HYPHEN).upper()


def _normalize_cell(value: object) -> str:
    """Normalize a worksheet cell value into a stripped string."""

    if value is None:
        return EMPTY
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _ensure_workbook_exists() -> None:
    """Create the workbook on disk if it is missing."""

    workbook_path = _workbook_path()
    base_dir = os.path.dirname(workbook_path)
    if not os.path.isdir(base_dir):
        os.makedirs(base_dir, exist_ok=True)
    if os.path.exists(workbook_path):
        return
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in _excel_sheets().values():
        sheet = workbook.create_sheet(title=sheet_name)
        if sheet_name == ENTRY_SHEET:
            sheet.append(ENTRY_HEADERS)
    try:
        workbook.save(workbook_path)
    except Exception as exc:  # pylint: disable=broad-except
        messagebox.showerror(ERROR_TITLE, localization.LIST_CREATE_FAILED_MSG.format(error=exc))
        log_error_loc("excel_create_failed", error=exc)


def _load_workbook():
    """Load the shared workbook, ensuring it exists first."""

    _ensure_workbook_exists()
    return load_workbook(_workbook_path())


def _load_workbook_readonly():
    """Load the shared workbook in read-only mode for startup cache reads."""

    _ensure_workbook_exists()
    return load_workbook(_workbook_path(), read_only=True, data_only=True)


def _entry_header_map(sheet, *, ensure_missing: bool = False) -> tuple[dict[str, int], bool]:
    """Return a map of entry-sheet header names to 1-based column indices."""

    changed = False
    if sheet.max_row < 1:
        sheet.append(ENTRY_HEADERS)
        changed = True
    header_map: dict[str, int] = {}
    first_row = next(sheet.iter_rows(min_row=1, max_row=1), ())
    for idx, cell in enumerate(first_row, start=1):
        header = _normalize_cell(cell.value).upper()
        if header:
            header_map[header] = idx
    if ensure_missing:
        next_col = max(header_map.values(), default=0) + 1
        for header in ENTRY_HEADERS:
            if header not in header_map:
                sheet.cell(row=1, column=next_col, value=header)
                header_map[header] = next_col
                next_col += 1
                changed = True
    return header_map, changed


def _entry_row_value(row, header_map: dict[str, int], header: str) -> str:
    """Return a normalized row value using the dynamic header map."""

    col_idx = header_map.get(header)
    if not col_idx or col_idx > len(row):
        return EMPTY
    return _normalize_cell(row[col_idx - 1].value)


def _set_row_value(row, header_map: dict[str, int], header: str, value: str) -> None:
    """Write a value into a row using the dynamic header map."""

    col_idx = header_map.get(header)
    if not col_idx:
        return
    row[col_idx - 1].value = value


def _normalize_extra_value(extra: str) -> str:
    """Normalize the extra flag saved into the entries sheet."""

    extra_raw = str(extra).strip()
    if not extra_raw or extra_raw.upper() in {NO_LED_VALUE, NO_LED_VALUE.upper()}:
        return NO_LED_VALUE
    return extra_raw.replace(UNDERSCORE, HYPHEN).upper()


def _build_entry_payload(
    ean: str,
    name: str,
    furniture_type: str,
    model: str,
    color1: str,
    color2: str,
    color3: str,
    extra_value: str,
    product_id: str = EMPTY,
) -> dict[str, str]:
    """Build a normalized entry payload used by the workbook and GUI cache."""

    return {
        EAN_HEADER: str(ean).strip(),
        NAME_HEADER: str(name).strip().upper(),
        TYPE_HEADER: str(furniture_type).strip().upper(),
        MODEL_HEADER: str(model).strip().upper(),
        COLOR1_HEADER: str(color1).strip().upper(),
        COLOR2_HEADER: str(color2).strip().upper() if color2 else EMPTY,
        COLOR3_HEADER: str(color3).strip().upper() if color3 else EMPTY,
        EXTRA_HEADER: _normalize_extra_value(extra_value),
        PRODUCT_ID_HEADER: str(product_id or EMPTY).strip().upper(),
    }


def _generate_product_id(existing_ids: set[str]) -> str:
    """Return a unique, stable product identifier for new or backfilled rows."""

    while True:
        candidate = f"PRD-{uuid.uuid4().hex[:12].upper()}"
        if candidate not in existing_ids:
            return candidate


def _write_entry_row(row, header_map: dict[str, int], payload: dict[str, str]) -> None:
    """Persist the normalized payload into the target worksheet row."""

    for header, value in payload.items():
        _set_row_value(row, header_map, header, value)


def _find_ean_conflict_row(sheet, header_map: dict[str, int], ean: str, skip_row=None):
    """Return a row using the same EAN, excluding the provided target row."""

    normalized = str(ean or EMPTY).strip().upper()
    if not normalized or normalized == NO_EAN_PLACEHOLDER:
        return None
    skip_row_number = None
    if skip_row:
        try:
            skip_row_number = skip_row[0].row
        except Exception:
            skip_row_number = None
    for row in sheet.iter_rows(min_row=2):
        try:
            row_number = row[0].row
        except Exception:
            row_number = None
        if skip_row_number is not None and row_number == skip_row_number:
            continue
        raw_ean = _entry_row_value(row, header_map, EAN_HEADER).upper()
        if raw_ean == normalized:
            return row
    return None


def prepare_excel_lists() -> Dict[str, Dict[str, dict] | list]:
    """Load all Excel lists into memory."""

    workbook = _load_workbook_readonly()
    lists: Dict[str, Dict[str, dict] | list] = {}
    try:
        for sheet_name in _excel_sheets().values():
            sheet = workbook[sheet_name]
            if sheet_name == ENTRY_SHEET:
                entries: Dict[str, dict] = {}
                records: list[dict[str, str]] = []
                header_map, _changed = _entry_header_map(sheet, ensure_missing=False)
                for row in sheet.iter_rows(min_row=2):
                    ean = _entry_row_value(row, header_map, EAN_HEADER)
                    if not ean:
                        continue
                    entry = _build_entry_payload(
                        ean,
                        _entry_row_value(row, header_map, NAME_HEADER),
                        _entry_row_value(row, header_map, TYPE_HEADER),
                        _entry_row_value(row, header_map, MODEL_HEADER),
                        _entry_row_value(row, header_map, COLOR1_HEADER),
                        _entry_row_value(row, header_map, COLOR2_HEADER),
                        _entry_row_value(row, header_map, COLOR3_HEADER),
                        _entry_row_value(row, header_map, EXTRA_HEADER),
                        _entry_row_value(row, header_map, PRODUCT_ID_HEADER),
                    )
                    entries[ean.strip()] = {
                        NAME_HEADER: entry[NAME_HEADER],
                        TYPE_HEADER: entry[TYPE_HEADER],
                        MODEL_HEADER: entry[MODEL_HEADER],
                        COLOR1_HEADER: entry[COLOR1_HEADER],
                        COLOR2_HEADER: entry[COLOR2_HEADER],
                        COLOR3_HEADER: entry[COLOR3_HEADER],
                        EXTRA_HEADER: entry[EXTRA_HEADER],
                        PRODUCT_ID_HEADER: entry[PRODUCT_ID_HEADER],
                    }
                    records.append(entry)
                lists[sheet_name] = entries
                lists[ENTRY_RECORDS_KEY] = records
            else:
                values: list[str] = []
                for row in sheet.iter_rows(
                    min_col=1,
                    max_col=1,
                    values_only=True,
                ):
                    cell_value = row[0]
                    if not cell_value:
                        continue
                    raw = str(cell_value).strip()
                    if sheet_name == EXTRAS_SHEET:
                        raw = raw.replace(UNDERSCORE, HYPHEN)
                    normalized = raw.upper()
                    if normalized not in values:
                        values.append(normalized)
                lists[sheet_name] = values
    finally:
        workbook.close()
    return lists


def _show_locked_dialog(reason: str) -> None:
    """Inform the user that the workbook is locked by another process."""

    title = LOCKED_TITLE
    text = localization.LANG.get(
        "excel_file_open",
        "Nie można zapisać listy. Plik Excel jest otwarty {reason}. Zamknij plik i spróbuj ponownie.",
    ).format(reason=reason)
    messagebox.showerror(title, text)


def _save_workbook(workbook: Workbook, error_event: str, **context) -> bool:
    """Persist the workbook and log a translated error message on failure."""

    try:
        workbook.save(_workbook_path())
        return True
    except Exception as exc:  # pylint: disable=broad-except
        messagebox.showerror(WRITE_ERROR_TITLE, localization.LIST_SAVE_FAILED_MSG.format(error=exc))
        context.setdefault("error", exc)
        log_error_loc(error_event, **context)
        return False


def add_to_list(sheet_name: str, value: str) -> bool:
    """Append a new normalised value to the given list sheet."""

    if not value:
        return False
    normalized = value.strip().upper()
    if sheet_name == _excel_sheets()[EXTRAS_SHEET]:
        normalized = normalized.replace(UNDERSCORE, HYPHEN)
    locked_by = get_file_lock_user(_workbook_path())
    if locked_by:
        reason = (
            LOCKED_BY_USER_TEMPLATE.format(user=locked_by)
            if isinstance(locked_by, str)
            else LOCKED_REASON_OTHER_PROCESS
        )
        _show_locked_dialog(reason)
        log_error_loc("excel_add_locked", value=normalized, list=sheet_name, reason=reason)
        return False
    workbook = _load_workbook()
    sheet = workbook[sheet_name]
    existing = {str(cell.value).strip().upper() for cell in sheet["A"] if cell.value}
    if normalized in existing:
        return True
    sheet.append([normalized])
    if _save_workbook(
        workbook,
        "excel_add_save_failed",
        value=normalized,
        list=sheet_name,
    ):
        log_info_loc("list_value_added", value=normalized, list=sheet_name)
        return True
    return False


def remove_from_list(sheet_name: str, value: str) -> None:
    """Remove the first occurrence of ``value`` from the target sheet."""

    locked_by = get_file_lock_user(_workbook_path())
    if locked_by:
        reason = (
            LOCKED_BY_USER_TEMPLATE.format(user=locked_by)
            if isinstance(locked_by, str)
            else LOCKED_REASON_OTHER_PROCESS
        )
        _show_locked_dialog(reason)
        log_error_loc("excel_remove_locked", value=value, list=sheet_name, reason=reason)
        return
    workbook = _load_workbook()
    sheet = workbook[sheet_name]
    removed = False
    for row in list(sheet.iter_rows(min_row=1)):
        cell = row[0]
        if cell.value and str(cell.value).strip().upper() == value.strip().upper():
            sheet.delete_rows(cell.row)
            removed = True
            break
    if removed and _save_workbook(
        workbook,
        "excel_remove_save_failed",
        value=value,
        list=sheet_name,
    ):
        log_info_loc("list_value_removed", value=value, list=sheet_name)


def _update_row(
    row,
    header_map: dict[str, int],
    name,
    furniture_type,
    model,
    color1,
    color2,
    color3,
    extra_value,
    product_id,
):
    """Mutate a worksheet row with the supplied normalised values."""

    payload = _build_entry_payload(
        _entry_row_value(row, header_map, EAN_HEADER),
        name,
        furniture_type,
        model,
        color1,
        color2,
        color3,
        extra_value,
        product_id,
    )
    _write_entry_row(row, header_map, payload)


def save_ean_entry(
    ean: str,
    name: str,
    furniture_type: str,
    model: str,
    color1: str,
    color2: str,
    color3: str,
    extra: str,
    product_id: str = EMPTY,
) -> dict[str, object] | bool:
    """Insert or update a row in the entries sheet, handling locks gracefully."""

    locked_by = get_file_lock_user(_workbook_path())
    if locked_by:
        reason = (
            LOCKED_BY_USER_TEMPLATE.format(user=locked_by)
            if isinstance(locked_by, str)
            else LOCKED_REASON_OTHER_PROCESS
        )
        messagebox = localization.O if hasattr(localization, "O") else None
        if messagebox:
            messagebox.showerror(
                LOCKED_TITLE,
                localization.LANG.get(
                    "excel_data_file_open",
                    "Nie można zapisać danych. Plik Excel jest otwarty {reason}. Zamknij plik i spróbuj ponownie.",
                ).format(reason=reason),
            )
        log_error_loc("excel_entry_save_locked", ean=ean, reason=reason)
        return False

    workbook = _load_workbook()
    sheet = workbook[_excel_sheets()[ENTRY_SHEET]]
    header_map, _changed = _entry_header_map(sheet, ensure_missing=True)

    payload = _build_entry_payload(
        ean,
        name,
        furniture_type,
        model,
        color1,
        color2,
        color3,
        extra,
        product_id,
    )
    trimmed = payload[EAN_HEADER]
    provided_product_id = payload[PRODUCT_ID_HEADER]
    existing_ids = {
        _entry_row_value(row, header_map, PRODUCT_ID_HEADER).upper()
        for row in sheet.iter_rows(min_row=2)
        if _entry_row_value(row, header_map, PRODUCT_ID_HEADER)
    }

    updated = False
    target_row = None
    for row in sheet.iter_rows(min_row=2):
        row_product_id = _entry_row_value(row, header_map, PRODUCT_ID_HEADER).upper()
        if provided_product_id and row_product_id == provided_product_id.upper():
            target_row = row
            updated = True
            break
    if target_row is None:
        for row in sheet.iter_rows(min_row=2):
            raw_ean = _entry_row_value(row, header_map, EAN_HEADER)
            if not raw_ean:
                continue
            if raw_ean.upper() == trimmed.upper():
                target_row = row
                updated = True
                break

    if target_row is not None:
        conflict_row = _find_ean_conflict_row(
            sheet,
            header_map,
            trimmed,
            skip_row=target_row,
        )
        if conflict_row is not None:
            conflict_product_id = _entry_row_value(
                conflict_row,
                header_map,
                PRODUCT_ID_HEADER,
            ).upper()
            messagebox.showwarning(
                localization.LANG.get("ean_duplicate_title", "Duplikat EAN"),
                localization.LANG.get(
                    "ean_duplicate_save_blocked",
                    "Kod EAN {ean} jest już zapisany w innym wpisie"
                    " (PRODUCT_ID: {product_id}).\n\n"
                    "Wczytaj istniejący wpis albo użyj innego EAN.",
                ).format(
                    ean=trimmed,
                    product_id=conflict_product_id or "BRAK-ID",
                ),
            )
            log_error_loc(
                "excel_entry_save_duplicate_ean",
                ean=trimmed,
                product_id=conflict_product_id or "BRAK-ID",
            )
            return False
        existing_product_id = _entry_row_value(
            target_row,
            header_map,
            PRODUCT_ID_HEADER,
        ).upper()
        final_product_id = provided_product_id or existing_product_id
        if not final_product_id:
            final_product_id = _generate_product_id(existing_ids)
            existing_ids.add(final_product_id)
        payload[PRODUCT_ID_HEADER] = final_product_id
        _write_entry_row(target_row, header_map, payload)
    else:
        if trimmed.upper() != NO_EAN_PLACEHOLDER:
            candidate = None
            for row in sheet.iter_rows(min_row=2):
                raw = _entry_row_value(row, header_map, EAN_HEADER).upper()
                if raw == NO_EAN_PLACEHOLDER:
                    existing = _build_entry_payload(
                        raw,
                        _entry_row_value(row, header_map, NAME_HEADER),
                        _entry_row_value(row, header_map, TYPE_HEADER),
                        _entry_row_value(row, header_map, MODEL_HEADER),
                        _entry_row_value(row, header_map, COLOR1_HEADER),
                        _entry_row_value(row, header_map, COLOR2_HEADER),
                        _entry_row_value(row, header_map, COLOR3_HEADER),
                        _entry_row_value(row, header_map, EXTRA_HEADER),
                        _entry_row_value(row, header_map, PRODUCT_ID_HEADER),
                    )
                    if (
                        existing[NAME_HEADER] == payload[NAME_HEADER]
                        and existing[TYPE_HEADER] == payload[TYPE_HEADER]
                        and existing[MODEL_HEADER] == payload[MODEL_HEADER]
                        and existing[COLOR1_HEADER] == payload[COLOR1_HEADER]
                        and existing[COLOR2_HEADER] == payload[COLOR2_HEADER]
                        and existing[COLOR3_HEADER] == payload[COLOR3_HEADER]
                        and existing[EXTRA_HEADER] == payload[EXTRA_HEADER]
                    ):
                        candidate = row
                        updated = True
                        break
            if candidate is not None:
                final_product_id = provided_product_id or _entry_row_value(
                    candidate,
                    header_map,
                    PRODUCT_ID_HEADER,
                ).upper()
                if not final_product_id:
                    final_product_id = _generate_product_id(existing_ids)
                    existing_ids.add(final_product_id)
                payload[PRODUCT_ID_HEADER] = final_product_id
                _write_entry_row(candidate, header_map, payload)
            else:
                final_product_id = provided_product_id or _generate_product_id(existing_ids)
                existing_ids.add(final_product_id)
                payload[PRODUCT_ID_HEADER] = final_product_id
                row_idx = sheet.max_row + 1
                for header, value in payload.items():
                    sheet.cell(row=row_idx, column=header_map[header], value=value)
        else:
            final_product_id = provided_product_id or _generate_product_id(existing_ids)
            existing_ids.add(final_product_id)
            payload[PRODUCT_ID_HEADER] = final_product_id
            row_idx = sheet.max_row + 1
            for header, value in payload.items():
                sheet.cell(row=row_idx, column=header_map[header], value=value)

    if _save_workbook(workbook, "excel_entry_save_failed", ean=ean):
        action_key = (
            "excel_entry_saved_updated" if updated else "excel_entry_saved_added"
        )
        log_info_loc(action_key, ean=ean)
        return {
            "updated": updated,
            "product_id": payload[PRODUCT_ID_HEADER],
            "entry": payload,
        }
    return False


# Backwards compatibility with the original constant name expected by app.py
Aw = SLOT_LABELS
