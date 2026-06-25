"""Import legacy JSON/Excel files into a SQLite data store."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .common import SQL_AVAILABLE_COLUMNS_KEY, SQL_COLUMN_MAP_KEY, SLOT_DEFS_KEY
from .excel_utils import (
    COLOR1_HEADER,
    COLOR2_HEADER,
    COLOR3_HEADER,
    EAN_HEADER,
    ENTRY_RECORDS_KEY,
    EXTRA_HEADER,
    MODEL_HEADER,
    NAME_HEADER,
    PRODUCT_ID_HEADER,
    TYPE_HEADER,
)
from .sqlite_store import LIST_SHEETS, SqliteStore

ENTRY_SHEET = "ENTRIES"
ENTRY_HEADERS = [
    EAN_HEADER,
    NAME_HEADER,
    TYPE_HEADER,
    MODEL_HEADER,
    COLOR1_HEADER,
    COLOR2_HEADER,
    COLOR3_HEADER,
    EXTRA_HEADER,
    PRODUCT_ID_HEADER,
]


def _read_json(path: Path, fallback: Any) -> Any:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError, TypeError):
        return fallback
    return payload


def _cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _header_map(row) -> dict[str, int]:
    mapping = {}
    for index, cell in enumerate(row or [], start=1):
        text = _cell(getattr(cell, "value", cell)).upper()
        if text:
            mapping[text] = index
    return mapping


def _row_value(row, mapping: dict[str, int], header: str) -> str:
    index = mapping.get(header)
    if not index or index > len(row):
        return ""
    return _cell(row[index - 1].value)


def _read_workbook_payload(path: Path) -> dict[str, Any]:
    payload: dict[str, Any] = {sheet: [] for sheet in LIST_SHEETS}
    payload[ENTRY_RECORDS_KEY] = []
    if not path.exists():
        return payload
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in LIST_SHEETS:
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            values = []
            for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
                value = _cell(row[0]).upper()
                if value and value not in values:
                    values.append(value)
            payload[sheet_name] = values
        if ENTRY_SHEET in workbook.sheetnames:
            sheet = workbook[ENTRY_SHEET]
            first_row = next(sheet.iter_rows(min_row=1, max_row=1), ())
            mapping = _header_map(first_row)
            records = []
            for row in sheet.iter_rows(min_row=2):
                ean = _row_value(row, mapping, EAN_HEADER)
                if not ean:
                    continue
                records.append(
                    {
                        EAN_HEADER: ean,
                        NAME_HEADER: _row_value(row, mapping, NAME_HEADER),
                        TYPE_HEADER: _row_value(row, mapping, TYPE_HEADER),
                        MODEL_HEADER: _row_value(row, mapping, MODEL_HEADER),
                        COLOR1_HEADER: _row_value(row, mapping, COLOR1_HEADER),
                        COLOR2_HEADER: _row_value(row, mapping, COLOR2_HEADER),
                        COLOR3_HEADER: _row_value(row, mapping, COLOR3_HEADER),
                        EXTRA_HEADER: _row_value(row, mapping, EXTRA_HEADER),
                        PRODUCT_ID_HEADER: _row_value(row, mapping, PRODUCT_ID_HEADER),
                    }
                )
            payload[ENTRY_RECORDS_KEY] = records
    finally:
        workbook.close()
    return payload


def import_legacy_to_sqlite(legacy_dir: str, database_path: str) -> dict[str, Any]:
    """Import supported legacy files from ``legacy_dir`` into ``database_path``."""

    source = Path(legacy_dir)
    store = SqliteStore(database_path)
    store.initialize()

    raw_config = _read_json(source / "config.json", {})
    config_imported = isinstance(raw_config, dict) and bool(raw_config)
    if config_imported:
        store.save_config(raw_config)
        columns = raw_config.get(SQL_AVAILABLE_COLUMNS_KEY, [])
        if isinstance(columns, list):
            store.save_sql_columns(columns)
        slot_defs = raw_config.get(SLOT_DEFS_KEY, [])
        sql_map = raw_config.get(SQL_COLUMN_MAP_KEY, {})
        if isinstance(slot_defs, list) and isinstance(sql_map, dict):
            store.save_slots(slot_defs, sql_map)

    lists_payload = _read_workbook_payload(source / "lists.xlsx")
    store.save_lists(lists_payload)

    users = _read_json(source / "web_users.json", [])
    if not isinstance(users, list):
        users = []
    store.save_users([item for item in users if isinstance(item, dict)])

    history = _read_json(source / "web_history.json", [])
    if not isinstance(history, list):
        history = []
    store.save_history([item for item in history if isinstance(item, dict)])

    file_index = _read_json(source / "file_index.json", {})
    file_index_imported = isinstance(file_index, dict) and bool(file_index)
    if file_index_imported:
        store.save_file_index_cache(file_index)

    records = lists_payload.get(ENTRY_RECORDS_KEY, [])
    return {
        "ok": True,
        "config": config_imported,
        "lists": sum(
            len(lists_payload.get(sheet, []))
            for sheet in LIST_SHEETS
            if isinstance(lists_payload.get(sheet, []), list)
        ),
        "entries": len(records) if isinstance(records, list) else 0,
        "users": len(users),
        "history": len(history),
        "file_index": file_index_imported,
    }
